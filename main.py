import requests
from bs4 import BeautifulSoup
import pandas as pd
import time
import logging
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class UniversityScraper:
    def __init__(self):
        self.base_url = "https://www.ox.ac.uk"
        self.headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        self.session = requests.Session()
        self.data = {
            'courses': [],
            'scholarships': []
        }
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    # Web scraping methods
    def get_courses(self):
        """Fetch undergraduate courses from Oxford"""
        try:
            url = "https://www.ox.ac.uk/admissions/undergraduate/courses/course-listing"
            response = self.session.get(url, headers=self.headers)
            soup = BeautifulSoup(response.text, 'html.parser')

            course_links = soup.select('a[href*="/admissions/undergraduate/courses/course-listing/"]')
            
            for link in course_links:
                course_name =  link.get_text(strip=True)
                if course_name and len(course_name) > 1:
                    self.data['courses'].append({'Course Name': course_name})
                    self.logger.info(f"Found course: {course_name}")

        except Exception as e:
            self.logger.error(f"Error in get_courses: {str(e)}")

    def get_scholarships(self):
        """Fetch scholarships from Oxford"""
        try:
            urls = [
                "https://www.ox.ac.uk/admissions/undergraduate/fees-and-funding/oxford-support",
                "https://www.ox.ac.uk/admissions/graduate/fees-and-funding/fees-funding-and-scholarship-search"
            ]
            
            for url in urls:
                response = self.session.get(url, headers=self.headers)
                soup = BeautifulSoup(response.text, 'html.parser')
                
                scholarship_sections = soup.find_all(['div', 'article'], class_=['scholarship-item', 'content-item'])
                
                if not scholarship_sections:
                    scholarship_sections = soup.find_all(['h2', 'h3', 'h4'])
                
                for section in scholarship_sections:
                    try:
                        name = section.text.strip()
                        description = ""
                        next_elem = section.find_next(['p', 'div'])
                        if next_elem:
                            description = next_elem.text.strip()
                        
                        eligibility = description if "eligible" in description.lower() else ""
                        amount = description if "£" in description or "$" in description else ""
                        
                        scholarship_data = {
                            'Scholarship Name': name,
                            'Description': description,
                            'Eligibility': eligibility,
                            'Amount': amount,
                            'Source URL': url
                        }
                        
                        self.data['scholarships'].append(scholarship_data)
                        self.logger.info(f"Processed scholarship: {name}")
                        
                    except Exception as e:
                        self.logger.error(f"Error processing scholarship section: {str(e)}")
                        continue
                
                time.sleep(2)  # Respectful delay between pages

        except Exception as e:
            self.logger.error(f"Error in get_scholarships: {str(e)}")

    # Data cleaning methods
    def clean_data(self):
        """Clean and format the scraped data"""
        self.clean_courses()
        self.clean_scholarships()

    def clean_courses(self):
        """Clean and format course data"""
        cleaned_courses = []
        for course in self.data['courses']:
            name = course['Course Name']
            name = ' '.join(name.split()).title()
            name = name.replace('Course Details', '').strip()
            cleaned_courses.append({'Course Name': name})
        
        # Remove duplicates while preserving order
        seen = set()
        self.data['courses'] = []
        for course in cleaned_courses:
            if course['Course Name'] not in seen:
                seen.add(course['Course Name'])
                self.data['courses'].append(course)

    def clean_scholarships(self):
        """Clean and format scholarship data"""
        for scholarship in self.data['scholarships']:
            for key in scholarship:
                if isinstance(scholarship[key], str):
                    scholarship[key] = ' '.join(scholarship[key].split())
                    if key == 'Scholarship Name':
                        scholarship[key] = scholarship[key].title()
                    if key == 'Amount' and scholarship[key]:
                        if '£' in scholarship[key] or '$' in scholarship[key]:
                            scholarship[key] = scholarship[key].replace(',', '')
                            scholarship[key] = scholarship[key].split()[0]

    # Excel formatting methods
    def format_excel_sheet(self, worksheet, df):
        """Apply formatting to Excel worksheet"""
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        regular_font = Font(size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Format headers
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = column
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Auto-adjust column width
            max_length = max(len(str(cell.value)), df[column].astype(str).apply(len).max())
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        # Format data cells
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = regular_font
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.border = border

        # Freeze the header row
        worksheet.freeze_panes = 'A2'

    def format_courses_sheet(self, worksheet):
        """Apply special formatting to courses worksheet"""
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12, name='Arial')
        course_font = Font(size=11, name='Arial')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

        # Format header
        header_cell = worksheet.cell(row=1, column=1)
        header_cell.value = "Course Name"
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        header_cell.border = border

        # Set column width
        worksheet.column_dimensions['A'].width = 60

        # Format course names
        for row_num, course in enumerate(self.data['courses'], start=2):
            cell = worksheet.cell(row=row_num, column=1)
            cell.value = course['Course Name']
            cell.font = course_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = border
            
            # Set row height
            worksheet.row_dimensions[row_num].height = 25

        # Freeze header row
        worksheet.freeze_panes = 'A2'

    # Data export method
    def export_to_excel(self, filename='oxford_university_data.xlsx'):
        """Export data to a well-formatted Excel file"""
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Courses sheet
                df_courses = pd.DataFrame({'Course Name': [course['Course Name'] for course in self.data['courses']]})
                df_courses.to_excel(writer, sheet_name='Courses', index=False)
                self.format_courses_sheet(writer.sheets['Courses'])

                # Scholarships sheet
                scholarship_columns = ['Scholarship Name', 'Amount', 'Eligibility', 'Description']
                df_scholarships = pd.DataFrame(self.data['scholarships'])
                if not df_scholarships.empty:
                    existing_columns = [col for col in scholarship_columns if col in df_scholarships.columns]
                    df_scholarships = df_scholarships[existing_columns]
                    df_scholarships.to_excel(writer, sheet_name='Scholarships', index=False)
                    self.format_excel_sheet(writer.sheets['Scholarships'], df_scholarships)

                # Summary sheet
                summary_data = {
                    'Category': ['Total Courses', 'Total Scholarships', 'Last Updated'],
                    'Value': [
                        len(self.data['courses']),
                        len(self.data['scholarships']),
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ]
                }
                df_summary = pd.DataFrame(summary_data)
                df_summary.to_excel(writer, sheet_name='Summary', index=False)
                self.format_excel_sheet(writer.sheets['Summary'], df_summary)

            self.logger.info(f"Data exported to {filename}")
            print(f"\nExport Summary:")
            print(f"Courses found: {len(self.data['courses'])}")
            print(f"Scholarships found: {len(self.data['scholarships'])}")
            print(f"Data has been exported to {filename}")

        except Exception as e:
            self.logger.error(f"Error exporting to Excel: {str(e)}")
            raise

    # Main execution method
    def scrape_all(self):
        """Run the complete scraping process"""
        self.logger.info("Starting web scraping process...")
        
        self.get_courses()
        self.logger.info(f"Found {len(self.data['courses'])} courses")
        
        time.sleep(2)
        
        self.get_scholarships()
        self.logger.info(f"Found {len(self.data['scholarships'])} scholarships")
        
        self.clean_data()
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.export_to_excel(f'oxford_data_{timestamp}.xlsx')
        
        self.logger.info("Web scraping completed!")

# Run the scraper
if __name__ == "__main__":
    scraper = UniversityScraper()
    scraper.scrape_all()